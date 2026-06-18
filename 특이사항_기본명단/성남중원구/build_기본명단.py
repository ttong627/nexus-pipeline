# -*- coding: utf-8 -*-
"""
경기도 성남시 중원구 기본명단 빌드 스크립트
- 입력: 기준명단.xlsx (마스터 10,372명, 주소 중복 多) + 명단관리.xlsx (실배송 4,873명, 기사·배송순번 권위)
- 처리: (이름+휴대폰) 기준 1인 1행 정제, 주소 표준형 정제, 명단관리 기사·배송순번 우선 반영
- 출력: 경기도 성남시 중원구 기본명단.xlsx
실행: python build_기본명단.py
"""
import os
import re
import sys
from collections import Counter, defaultdict

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
SRC_REF = os.path.join(BASE_DIR, "기준명단.xlsx")        # 기준명단
SRC_MNG = os.path.join(BASE_DIR, "명단관리.xlsx")        # 명단관리
OUT_PATH = os.path.join(BASE_DIR, "경기도 성남시 중원구 기본명단.xlsx")

ADMIN_PREFIX = re.compile(r"^경기도?성남시중원구")


def load_rows(path):
    """헤더(1~2행) 제외, 비어있지 않은 데이터 행만 반환."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = [r for r in ws.iter_rows(min_row=3, values_only=True)
            if any(c not in (None, "") for c in r)]
    wb.close()
    return rows


def s(x):
    """어떤 타입이든 안전하게 trim된 문자열로."""
    if x is None:
        return ""
    return str(x).strip()


def norm_phone(p):
    if p is None:
        return ""
    return re.sub(r"\D", "", str(p))


def fmt_phone(p):
    """숫자만 추출 후 010-xxxx-xxxx 형태로. 형식 불명확하면 원본 숫자 유지."""
    d = norm_phone(p)
    if len(d) == 11:
        return f"{d[:3]}-{d[3:7]}-{d[7:]}"
    if len(d) == 10:
        return f"{d[:3]}-{d[3:6]}-{d[6:]}"
    return d


_HO_TOKEN = r"[0-9A-Za-z가-힣층\-]+호"      # 201호 · 1층102호 · B01호 · 지하층B03호
_DONG_TOKEN = r"\d+동"                       # 101동


_PAREN_RE = re.compile(r"\([^()]*\)")
_POST_PAREN_DUP = re.compile(rf"({_HO_TOKEN})(\([^()]*\)),?\1", re.IGNORECASE)


def dedupe_detail(s):
    """원본 오타로 상세주소(호/동)·법정동 괄호가 두 번 들어간 경우만 1개로 정리.
    서로 다른 토큰(101동 502호)·다른 괄호((옥탑)(중앙동))는 절대 건드리지 않음."""
    # 1) 붙어버린 동일 토큰 반복: 201호201호 -> 201호
    for pat in (rf"({_HO_TOKEN})\1", rf"({_DONG_TOKEN})\1"):
        prev = None
        while prev != s:
            prev = s
            s = re.sub(pat, r"\1", s)
    # 2) 괄호 뒤(또는 콤마)로 이어진 동일 호 토큰 제거:
    #    302호(하대원동,강남아트빌라)302호 -> 302호(하대원동,강남아트빌라)
    prev = None
    while prev != s:
        prev = s
        s = _POST_PAREN_DUP.sub(r"\1\2", s)
    # 3) 내용이 완전히 같은 괄호 중복 제거(뒤쪽 1개만 유지): (금광동)…(금광동) -> …(금광동)
    cnt = Counter(_PAREN_RE.findall(s))
    for p, c in cnt.items():
        for _ in range(c - 1):
            s = s.replace(p, "", 1)        # 앞쪽부터 c-1개 제거
    # 4) 콤마 분리 순수 호/동 토큰 중복 제거
    parts = s.split(",")
    out = []
    seen_pure = set()
    for p in parts:
        pt = p.strip()
        is_pure = bool(re.fullmatch(_HO_TOKEN, pt) or re.fullmatch(_DONG_TOKEN, pt))
        if is_pure:
            # (a) 동일 순수 토큰이 더 긴 다른 파트에 포함됨(예: '201호' vs '201호(금광동)') → 제거
            if any(pt in o.strip() and o.strip() != pt for o in parts):
                continue
            # (b) 이미 같은 순수 토큰을 담았음 → 중복 제거(첫 1개만 유지)
            if pt in seen_pure:
                continue
            seen_pure.add(pt)
        out.append(p)
    return ",".join(out)


def clean_addr(a):
    """명단관리 표준형으로 정제: 공백 제거 · 행정구역 접두어 제거 · 상세중복 제거 · 콤마 정리."""
    if a is None:
        return ""
    s = str(a).strip()
    s = re.sub(r"\s+", "", s)          # 내부 공백 전부 제거
    s = ADMIN_PREFIX.sub("", s)        # '경기도성남시중원구' 접두어 제거(별도 칼럼 존재)
    s = dedupe_detail(s)               # 상세(호/동) 동일 토큰 중복 제거 (원본 오타 대응)
    s = re.sub(r",{2,}", ",", s)       # 중복 콤마 정리
    s = s.strip(",")                   # 양끝 콤마 제거
    return s


def addr_score(a):
    """주소 정보량 점수 — 중복 변형 중 가장 풍부한 주소 선택용."""
    c = clean_addr(a)
    score = len(c)
    if "호" in c:
        score += 30
    if "(" in c:
        score += 20
    if re.search(r"\d", c):
        score += 10
    return score


def seq_value(v):
    """배송순번 숫자화. 정렬용(미배정/None은 맨 뒤)."""
    try:
        return int(v)
    except (TypeError, ValueError):
        return None


def main():
    ref = load_rows(SRC_REF)
    # 기준명단 cols: 0순번 1광역시도 2시군구 3읍면동 4수령자명 5휴대폰 6주소 7기사 8특이사항 9배송순번
    mng = load_rows(SRC_MNG)
    # 명단관리 cols: 0순번 1취약계층 2배송월 3광역시도 4시군구 5읍면동 6배송기관명 7배송기사
    #                8성명 9포수 10주소 11휴대폰 12전화번호 13품목 14배송순번 15생년월일 16특이사항 17가구원수 18문자수신여부

    # --- 명단관리 lookup (권위 데이터) ---
    mlut = {}
    for r in mng:
        k = ((r[8] or "").strip(), norm_phone(r[11]))
        if not k[0]:
            continue
        mlut[k] = {
            "기사": s(r[7]),
            "배송순번": r[14],
            "주소": r[10],
            "취약계층": s(r[1]),
            "품목": s(r[13]),
            "특이사항": s(r[16]),
        }

    # --- 기준명단 (이름+휴대폰) 그룹핑 ---
    grp = defaultdict(list)
    for r in ref:
        k = ((r[4] or "").strip(), norm_phone(r[5]))
        if not k[0]:
            continue
        grp[k].append(r)

    records = []
    n_overlap = n_refonly = 0
    for k, rows in grp.items():
        best = max(rows, key=lambda r: addr_score(r[6]))   # 정보량 최다 주소 행
        name, phone_d = k
        sido = s(best[1])
        sigungu = s(best[2])
        emd = s(best[3])

        in_m = k in mlut
        if in_m:
            n_overlap += 1
            m = mlut[k]
            # 주소: 기준명단 best vs 명단관리 중 정보량 많은 쪽
            addr = clean_addr(max([best[6], m["주소"]], key=addr_score))
            gisa = m["기사"] or s(best[7])
            seq = m["배송순번"]
            note = m["특이사항"] or s(best[8])
            vuln = m["취약계층"]
            item = m["품목"]
            src = "명단관리"
        else:
            n_refonly += 1
            addr = clean_addr(best[6])
            # 기준명단 전용: 변형마다 기사가 충돌(58%)해 현재값 신뢰 불가 → 비움. 명단관리 갱신 시 채워짐.
            gisa = ""
            seq = None
            note = s(best[8])
            vuln = ""
            item = ""
            src = "기준명단"

        records.append({
            "광역시도": sido, "시군구": sigungu, "읍면동": emd,
            "수령자명": name, "휴대폰": fmt_phone(phone_d), "주소": addr,
            "기사": gisa, "배송순번": seq, "취약계층": vuln, "품목": item,
            "특이사항": note, "데이터출처": src,
        })

    # --- 정렬: 명단관리(실배송) 먼저 → 기사 → 배송순번 / 기준명단 전용은 뒤(읍면동·기사순) ---
    def sort_key(rec):
        s = seq_value(rec["배송순번"])
        is_mng = 0 if rec["데이터출처"] == "명단관리" else 1
        if is_mng == 0:
            return (0, rec["기사"], (s is None, s if s is not None else 0), rec["읍면동"])
        return (1, rec["읍면동"], rec["기사"], (s is None, s if s is not None else 0))
    records.sort(key=sort_key)

    # --- 출력 워크북 ---
    headers = ["순번", "광역시도", "시군구", "읍면동", "수령자명", "휴대폰", "주소",
               "기사", "배송순번", "취약계층", "품목", "특이사항", "데이터출처"]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "기본명단"

    title_font = Font(name="맑은 고딕", size=14, bold=True, color="FFFFFF")
    hdr_font = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
    cell_font = Font(name="맑은 고딕", size=10)
    title_fill = PatternFill("solid", fgColor="1F4E78")
    hdr_fill = PatternFill("solid", fgColor="2E75B6")
    mng_fill = PatternFill("solid", fgColor="FFF2CC")   # 명단관리 출처 행 강조
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    # 1행: 제목
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    tcell = ws.cell(row=1, column=1, value="경기도 성남시 중원구 기본명단")
    tcell.font = title_font
    tcell.fill = title_fill
    tcell.alignment = center
    ws.row_dimensions[1].height = 26

    # 2행: 헤더
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[2].height = 20

    # 3행~: 데이터
    for i, rec in enumerate(records, 1):
        row = 2 + i
        seq = seq_value(rec["배송순번"])
        # 명단관리 미포함자의 미배정(0/None) 순번은 빈칸으로 표기
        if rec["데이터출처"] != "명단관리" and (seq is None or seq == 0):
            seq_out = ""
        else:
            seq_out = seq if seq is not None else ""
        vals = [i, rec["광역시도"], rec["시군구"], rec["읍면동"], rec["수령자명"],
                rec["휴대폰"], rec["주소"], rec["기사"],
                seq_out, rec["취약계층"], rec["품목"],
                rec["특이사항"], rec["데이터출처"]]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.font = cell_font
            cell.border = border
            cell.alignment = left if c in (5, 7, 12) else center
            if c == 6:
                cell.number_format = "@"   # 휴대폰 텍스트 고정
        if rec["데이터출처"] == "명단관리":
            for c in range(1, len(headers) + 1):
                ws.cell(row=row, column=c).fill = mng_fill

    # 열 너비
    widths = [6, 8, 12, 10, 10, 14, 38, 9, 9, 10, 10, 24, 10]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}{2 + len(records)}"

    wb.save(OUT_PATH)

    print("=" * 60)
    print("[완료] 저장:", OUT_PATH)
    print(f"  총 인원      : {len(records):,}명")
    print(f"  명단관리 반영 : {n_overlap:,}명 (기사·배송순번 우선)")
    print(f"  기준명단 전용 : {n_refonly:,}명")
    print(f"  원본 기준명단 행: {len(ref):,} → 정제 후 {len(grp):,}명 "
          f"(중복 {len(ref) - len(grp):,}행 제거)")


if __name__ == "__main__":
    main()
