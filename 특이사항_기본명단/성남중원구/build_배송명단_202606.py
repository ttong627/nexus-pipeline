# -*- coding: utf-8 -*-
"""
2026년 6월 성남중원구 정부양곡 배송명단 생성 + 기사배정
- 입력: (중원구)2026년 6월 정부양곡 택배요청 명단(2026).xlsx  (수급자+차상위 시트)
- 기사배정 규칙(4월 명단관리 + 수정규칙):
  1) 단독 전담 동(9개)  → 그 동의 4월 담당 기사
  2) 은행2동(도로 분할) → 도로별 4월 다수 기사 (한 도로=한 기사, 혼재 자동정리)
  3) 도촌동(대단지)     → 아파트 동(棟)번호 구간 (저층 허인구 / 고층 권순갑)
  4) 규칙 불가 시       → 4월 (이름+휴대폰) 승계, 그래도 없으면 확인필요
실행: python build_배송명단_202606.py
"""
import os
import re
import sys
import importlib.util
from collections import Counter, defaultdict

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
JUNE = os.path.join(BASE_DIR, "(중원구)2026년 6월 정부양곡 택배요청 명단(2026).xlsx")
APRIL = os.path.join(BASE_DIR, "명단관리.xlsx")
OUT = os.path.join(BASE_DIR, "2026년 6월 성남중원구 배송명단.xlsx")

# build_기본명단.py 의 주소 정제 함수 재사용
_spec = importlib.util.spec_from_file_location("bld", os.path.join(BASE_DIR, "build_기본명단.py"))
_bld = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(_bld)
clean_addr = _bld.clean_addr


def npn(p):
    return re.sub(r"\D", "", str(p)) if p else ""


def s(x):
    return "" if x is None else str(x).strip()


def roadkey(a):
    a = str(a or "").replace(" ", "")
    m = re.match(r"([가-힣A-Za-z]+로(?:\d+번길)?)", a)
    return m.group(1) if m else a.split(",")[0][:10]


def dong_no(a):
    """아파트 동(棟) 번호 추출: '204동' 또는 '204-502호' 형태."""
    a = str(a or "").replace(" ", "")
    m = re.search(r"(\d{2,3})동", a) or re.search(r"(\d{2,3})-\d+호", a)
    return int(m.group(1)) if m else None


def road_no(a):
    a = str(a or "").replace(" ", "")
    m = re.search(r"도촌남로(\d+)", a)
    return int(m.group(1)) if m else None


# ── 4월(명단관리)에서 배정 규칙 학습 ─────────────────────────────
def learn_april():
    wb = openpyxl.load_workbook(APRIL, data_only=True)
    ws = wb.active
    rows = [r for r in ws.iter_rows(min_row=3, values_only=True) if any(c not in (None, "") for c in r)]
    wb.close()
    # 5읍면동 7배송기사 8성명 10주소 11휴대폰
    person, dong_drv = {}, defaultdict(Counter)
    eunhaeng_road = defaultdict(Counter)
    for r in rows:
        name, ph = s(r[8]), npn(r[11])
        drv, dong, addr = s(r[7]), s(r[5]), s(r[10])
        if name:
            person[(name, ph)] = drv
        dong_drv[dong][drv] += 1
        if dong == "은행2동":
            eunhaeng_road[roadkey(addr)][drv] += 1
    single = {d: c.most_common(1)[0][0] for d, c in dong_drv.items() if len(c) == 1}
    eunhaeng = {rd: c.most_common(1)[0][0] for rd, c in eunhaeng_road.items()}
    dong_major = {d: c.most_common(1)[0][0] for d, c in dong_drv.items()}
    return person, single, eunhaeng, dong_major


def assign_driver(dong, addr, name, ph, ctx):
    person, single, eunhaeng, dong_major = ctx
    if dong in single:
        return single[dong], "단독전담동"
    if dong == "은행2동":
        rd = roadkey(addr)
        if rd in eunhaeng:
            return eunhaeng[rd], "도로규칙"
        if (name, ph) in person:
            return person[(name, ph)], "4월승계"
        return dong_major.get(dong, ""), "동최다"
    if dong == "도촌동":
        dn = dong_no(addr)
        if dn is not None:
            return ("허인구" if dn < 700 else "권순갑"), "동번호규칙"
        rn = road_no(addr)
        if rn is not None:
            return ("허인구" if rn < 100 else "권순갑"), "도로번호규칙"
        if (name, ph) in person:
            return person[(name, ph)], "4월승계"
        return dong_major.get(dong, ""), "동최다"
    if (name, ph) in person:
        return person[(name, ph)], "4월승계"
    return "", "확인필요"


# ── 6월 명단 파싱 ────────────────────────────────────────────────
def parse_june():
    wb = openpyxl.load_workbook(JUNE, data_only=True)
    out = []
    for sheet, vuln_default in (("수급자", "기초수급자"), ("차상위", "차상위")):
        ws = wb[sheet]
        # 2연번 ... 컬럼: 1구분 2성명 4동 5상세주소 6일반전화 7휴대전화 8수량포 10가구원수 11비고
        for r in ws.iter_rows(min_row=5, values_only=True):
            name = s(r[2])
            if not name or "세대" in name or name == "성명":
                continue
            addr = s(r[5])
            if not addr:
                continue
            mobile = npn(r[7]) or npn(r[6])
            try:
                po = int(r[8]) if r[8] not in (None, "") else 1
            except (ValueError, TypeError):
                po = 1
            try:
                gagu = int(r[10]) if r[10] not in (None, "") else ""
            except (ValueError, TypeError):
                gagu = ""
            out.append({
                "구분": s(r[1]) or vuln_default,
                "동": s(r[4]),
                "성명": name,
                "휴대폰": mobile,
                "주소": clean_addr(addr),
                "포수": po,
                "가구원수": gagu,
                "비고": s(r[11]),
            })
    wb.close()
    return out


def fmt_phone(d):
    if len(d) == 11:
        return f"{d[:3]}-{d[3:7]}-{d[7:]}"
    if len(d) == 10:
        return f"{d[:3]}-{d[3:6]}-{d[6:]}"
    return d


def main():
    ctx = learn_april()
    recs = parse_june()

    for rec in recs:
        drv, why = assign_driver(rec["동"], rec["주소"], rec["성명"], rec["휴대폰"], ctx)
        rec["기사"] = drv
        rec["배정근거"] = why

    # 배송순번: 기사별 (동 → 도로 → 번지/동번호) 정렬 후 1..N
    def seqkey(rec):
        a = rec["주소"].replace(" ", "")
        rno = re.search(r"(\d+)", roadkey(a) + "0")
        return (rec["동"], roadkey(a), dong_no(a) or 0, a)
    by_drv = defaultdict(list)
    for rec in recs:
        by_drv[rec["기사"]].append(rec)
    for drv, lst in by_drv.items():
        lst.sort(key=seqkey)
        for i, rec in enumerate(lst, 1):
            rec["배송순번"] = i if drv else ""

    # 정렬: 동 → 기사 → 배송순번
    recs.sort(key=lambda r: (r["동"], r["기사"], r["배송순번"] if isinstance(r["배송순번"], int) else 0))

    # ── 출력 ──
    headers = ["순번", "구분", "읍면동", "세대주명", "휴대폰", "주소", "포수",
               "가구원수", "기사", "배송순번", "비고", "배정근거"]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "6월배송명단"
    tfont = Font(name="맑은 고딕", size=14, bold=True, color="FFFFFF")
    hfont = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
    cfont = Font(name="맑은 고딕", size=10)
    tfill = PatternFill("solid", fgColor="1F4E78")
    hfill = PatternFill("solid", fgColor="2E75B6")
    chk = PatternFill("solid", fgColor="FFC7CE")   # 확인필요 강조
    thin = Side(style="thin", color="BFBFBF")
    bd = Border(left=thin, right=thin, top=thin, bottom=thin)
    cen = Alignment(horizontal="center", vertical="center")
    lft = Alignment(horizontal="left", vertical="center")

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    t = ws.cell(row=1, column=1, value="2026년 6월 성남중원구 정부양곡 배송명단")
    t.font, t.fill, t.alignment = tfont, tfill, cen
    ws.row_dimensions[1].height = 26
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font, cell.fill, cell.alignment, cell.border = hfont, hfill, cen, bd

    for i, rec in enumerate(recs, 1):
        row = 2 + i
        vals = [i, rec["구분"], rec["동"], rec["성명"], fmt_phone(rec["휴대폰"]),
                rec["주소"], rec["포수"], rec["가구원수"], rec["기사"],
                rec["배송순번"], rec["비고"], rec["배정근거"]]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.font, cell.border = cfont, bd
            cell.alignment = lft if c in (4, 6, 11) else cen
            if c == 5:
                cell.number_format = "@"
        if rec["배정근거"] == "확인필요":
            for c in range(1, len(headers) + 1):
                ws.cell(row=row, column=c).fill = chk

    widths = [6, 14, 9, 10, 14, 40, 6, 8, 8, 9, 20, 11]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}{2 + len(recs)}"
    wb.save(OUT)

    # ── 리포트 ──
    tot_po = sum(r["포수"] for r in recs)
    print("=" * 60)
    print("[완료] 저장:", OUT)
    print(f"  총 세대: {len(recs):,} / 총 포수: {tot_po:,}")
    print("  기사별 (세대/포수):")
    drv_stat = defaultdict(lambda: [0, 0])
    for r in recs:
        drv_stat[r["기사"] or "(미배정)"][0] += 1
        drv_stat[r["기사"] or "(미배정)"][1] += r["포수"]
    for g, (c, p) in sorted(drv_stat.items(), key=lambda x: -x[1][1]):
        print(f"    {g:<8}: {c:>4}세대 / {p:>4}포")
    print("  배정근거 분포:", dict(Counter(r["배정근거"] for r in recs)))


if __name__ == "__main__":
    main()
